import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import MergedCell
import win32com.client as win32
import pandas as pd
# pip install -i https://pypi.tuna.tsinghua.edu.cn/simple pandas
# pip install -i https://pypi.tuna.tsinghua.edu.cn/simple openpyxl
# pip install -i https://pypi.tuna.tsinghua.edu.cn/simple xlwt
path = '.'

class Head:
    def __init__(self):
        self.recordNo=''  #提单号
        self.vesselName=''  #船名航次
        self.CNTRNO=''  #柜号
        self.sealNO=''  #封条号
        self.dischargingPort=''  #卸货港
        self.destinationPort = ''  #目的港

    def __str__(self):
        print("head:\nrecordNo", self.recordNo, "\nvesselName", self.vesselName, "\nCNTRNO", self.CNTRNO,
              "\nsealNO", self.sealNO, "\ndischargingPort", self.dischargingPort, "\ndestinationPort", self.destinationPort,)
        return ""

'''
用来转换xls与xlsx格式的
'''
def exchange(dir):
    '''
    :param dir: product_count,product_trend,product_before15 文件夹
    :return:
    '''
    path=os.path.dirname(os.path.abspath(__file__))
    path = os.path.join(path,dir)
    files = os.listdir(path)
    for file_name in files:
        if file_name.rsplit('.',1)[-1]=='xls':
            fname = os.path.join(path,file_name)
            excel = win32.gencache.EnsureDispatch('Excel.Application')
            wb = excel.Workbooks.Open(fname)
            #在原来的位置创建出：原名+'.xlsx'文件
            wb.SaveAs(fname+"x", FileFormat = 51)    #FileFormat = 51 is for .xlsx extension
            wb.Close()                               #FileFormat = 56 is for .xls extension
            excel.Application.Quit()
            os.remove(fname)

def parser_merged_cell(sheet: Worksheet, row, col):
    """
    检查是否为合并单元格并获取对应行列单元格的值。
    如果是合并单元格，则取合并区域左上角单元格的值作为当前单元格的值,否则直接返回该单元格的值
    :param sheet: 当前工作表对象
    :param row: 需要获取的单元格所在行
    :param col: 需要获取的单元格所在列
    :return: 
    """
    cell = sheet.cell(row=row, column=col)
    if isinstance(cell, MergedCell):  # 判断该单元格是否为合并单元格
        for merged_range in sheet.merged_cell_ranges:  # 循环查找该单元格所属的合并区域
            if cell.coordinate in merged_range:
                # 获取合并区域左上角的单元格作为该单元格的值返回
                cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                break
    return cell.value

def scan_file(path):
    filelist = list()
    all = os.listdir(path)
    for f in all:
        if f.startswith("排柜表") and f.endswith('.xlsx'):
            filelist.append(f)

    return filelist


def main():
    filelist = scan_file(path)
    print(filelist)

    exchange('.')

    for infile in filelist:
        op_file(infile)
    print('ok')


def read_xls(infile):
    wb = load_workbook(infile)
    ws = wb.active
    df = pd.read_excel(infile)
    df = df.rename(columns={
            'Unnamed: 1':'B',
            'Unnamed: 2':'C',
            'Unnamed: 3':'D',
            'Unnamed: 4':'E',
            'Unnamed: 5':'F',
            'Unnamed: 6':'G',
            'Unnamed: 7':'H',
            'Unnamed: 8':'I',
            'Unnamed: 9':'J',
            'Unnamed: 21':'V',
        })

    head = Head()
    print(df.head())
    head.recordNo = df['G'][14]
    head.vesselName = df['V'][14]
    head.CNTRNO = df['E'][8]
    print(head)

    # print(df['B'])

    start_row = 20
    end_row = start_row+1000  # 读处理1000行
    start_col = 2
    end_col = start_col+36  # 读 列


    rows = ws.iter_rows(min_col=start_col, max_col=end_col,
                        min_row=start_row, max_row=end_row)
    cells_value = list()
    # cells_value=[['a','b','c'],[1,2,3]]
    n_remark_col=12
    n_row = 0
    for cells in rows:
        n_row += 1
        row = [cell.value for cell in cells]
        print(row)
        print(n_row,row[n_remark_col])
        if n_row >1 :
            row[n_remark_col] = parser_merged_cell(ws, n_row+start_row -1, n_remark_col+start_col)
        print(row[n_remark_col])
        # print(row)
        if isinstance(row[0], str) and not row[0].isdigit() and n_row != 1:
            print('read row end with:', row[0], type(row[0]))
            break
        # print(row[0], type(row[0]))
        cells_value.append(row)
        # print(cells_value)

    # Load the list to a dataframe
    df = pd.DataFrame(cells_value)
    # Grab the first row for the header
    df_header = df.iloc[0]
    # Get the data except the 1st row
    df = df[1:]
    # Set the 1st row as header
    df.columns = df_header

    # Reset df index
    df.reset_index(drop=True, inplace=True)
    print(df.head)
    return df, head



def op_file(infile):
    # df = pd.read_csv(infile, sep=",", encoding="gbk")
    df, head = read_xls(infile)

    df = df[['工作号/入仓单号', '件数', '毛重', '体积', '报关', '排柜备注']]
    print(df)
    dfrank = df.sort_values(by=['报关','排柜备注','工作号/入仓单号'], ascending=[False,False,True])
    dfrank.insert(1, 'HS code', '')
    dfrank.insert(0, '关单号', '')
    dfrank.insert(0, '顺序', '')
    df = dfrank
    wb = Workbook()
    ws = wb.active
    recordNo = head.recordNo
    ws.append(["提单号", head.recordNo])
    ws.append(["船名航次", head.vesselName])
    ws.append(["柜号", head.CNTRNO])
    ws.append(["封条号", head.sealNO])
    ws.append(["卸货港", head.dischargingPort])
    ws.append(["目的港", head.destinationPort])

    head_row_count = 6
    content_row_count = 0

    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
        content_row_count += 1

    headfont = Font(name="微软雅黑", size=10, bold=False,
                    italic=False, color="FF0000")
    titlefont = Font(name="微软雅黑", size=10, bold=True,
                     italic=False, color="000000")
    contentfont = Font(name="微软雅黑", size=10, bold=False,
                       italic=False, color="000000")
    alignment = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    n_row = 0
    for row in ws.iter_rows():
        n_row += 1
        if n_row <= head_row_count:
            print(row)
            col = 0
            for cell in row:
                col += 1
                print(cell, cell.value, "headfont")
                cell.font = headfont
                cell.alignment = alignment
                cell.border = border
                if col >= 2:
                    break
            continue
        for cell in row:
            print(cell, cell.value, "contentfont")
            cell.font = contentfont
            cell.alignment = alignment
            cell.border = border

    for i in range(1, df.shape[1]+1):
        cell = ws.cell(row=head_row_count + 1, column=i)
        print(cell.value)
        cell.font = titlefont

    # 设置连续行行高：
    for r in range(head_row_count+1, head_row_count+content_row_count+1):  # 注意，行和列的序数，都是从1开始
        ws.row_dimensions[r].height = 22  #
    # 设置连续列列宽：
    for c in range(1, 8):  # 注意，列序数从1开始，但必须转变为A\B等字母
        w = get_column_letter(c)  # 把列序数转变为字母
        ws.column_dimensions[w].width = 16
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["I"].width = 50
    recordNo = recordNo.replace("/", " ")
    wb.save(recordNo + '-进港舱单.xlsx')


if __name__ == '__main__':
    main()
    # infile = '排柜表2022.xlsx'
    # read_xls(infile)
# 